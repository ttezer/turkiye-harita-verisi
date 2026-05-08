#!/usr/bin/env python

from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
UAB_PATH = ROOT / "source" / "reference" / "ilce-listesi.xlsx"
PROVINCES_PATH = ROOT / "data" / "normalized" / "provinces.metadata.partial.json"
DISTRICTS_PATH = ROOT / "data" / "normalized" / "districts.metadata.partial.json"
OUTPUT_PATH = ROOT / "source" / "reference" / "display-name-overrides.json"


TRANSLIT = str.maketrans(
    {
        "Ç": "C",
        "ç": "c",
        "Ğ": "G",
        "ğ": "g",
        "İ": "i",
        "I": "i",
        "ı": "i",
        "Ö": "O",
        "ö": "o",
        "Ş": "S",
        "ş": "s",
        "Ü": "U",
        "ü": "u",
    }
)

TR_LOWER = str.maketrans(
    {
        "I": "ı",
        "İ": "i",
        "Ç": "ç",
        "Ğ": "ğ",
        "Ö": "ö",
        "Ş": "ş",
        "Ü": "ü",
    }
)

TR_UPPER = str.maketrans(
    {
        "i": "İ",
        "ı": "I",
        "ç": "Ç",
        "ğ": "Ğ",
        "ö": "Ö",
        "ş": "Ş",
        "ü": "Ü",
    }
)


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def fold_turkish_to_ascii(value: str) -> str:
    return value.translate(TRANSLIT).strip().lower()


def normalize_key(value: str) -> str:
    return " ".join(fold_turkish_to_ascii(value).split())


def canonical_plate_code(value: str) -> str:
    return str(int(value)).zfill(2)


def tr_lower(value: str) -> str:
    return value.translate(TR_LOWER).lower()


def tr_upper_char(value: str) -> str:
    return value.translate(TR_UPPER).upper()


def tr_title_word(word: str) -> str:
    if not word:
        return word
    lowered = tr_lower(word)
    return tr_upper_char(lowered[0]) + lowered[1:]


def tr_title(value: str) -> str:
    words = []
    for word in value.replace("/", " / ").split():
        if word == "/":
            words.append(word)
            continue
        parts = [tr_title_word(part) for part in word.split("-")]
        words.append("-".join(parts))
    return " ".join(words).replace(" / ", "/")


def load_uab_rows() -> List[dict]:
    workbook = load_workbook(UAB_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    header_index = {name: idx for idx, name in enumerate(headers)}
    rows = []

    for values in worksheet.iter_rows(min_row=2, values_only=True):
        district_name = values[header_index["AD"]]
        province_code = values[header_index["IL_KODU"]]
        province_name = values[header_index["IL_ADI"]]

        if not district_name or not province_code or not province_name:
            continue

        rows.append(
            {
                "province_code": str(province_code).zfill(2),
                "province_name": str(province_name).strip(),
                "district_name": str(district_name).strip(),
            }
        )

    return rows


def main():
    if not UAB_PATH.exists():
        raise SystemExit(f"Missing reference file: {UAB_PATH}")

    provinces = read_json(PROVINCES_PATH)
    districts = read_json(DISTRICTS_PATH)
    uab_rows = load_uab_rows()
    provinces_by_source = {item["hdx_id"]: item for item in provinces}

    province_name_by_code: Dict[str, str] = {}
    province_buckets = defaultdict(set)

    for row in uab_rows:
        province_buckets[canonical_plate_code(row["province_code"])].add(tr_title(row["province_name"]))

    for province_code, names in province_buckets.items():
        if len(names) == 1:
            province_name_by_code[province_code] = next(iter(names))

    province_overrides = []
    province_name_map = {}
    for item in provinces:
        corrected_name = province_name_by_code.get(canonical_plate_code(item["plate_code"]))
        if corrected_name and corrected_name != item["name"]:
            province_overrides.append(
                {
                    "hdx_id": item["hdx_id"],
                    "plate_code": canonical_plate_code(item["plate_code"]),
                    "name": corrected_name,
                    "source": "uab_validation_only_exact_province_code",
                }
            )

        province_name_map[item["hdx_id"]] = corrected_name or item["name"]

    district_uab_name_map = defaultdict(set)
    for row in uab_rows:
        key = (canonical_plate_code(row["province_code"]), normalize_key(row["district_name"]))
        district_uab_name_map[key].add(tr_title(row["district_name"]))

    district_hdx_map = defaultdict(list)
    for item in districts:
        key = (canonical_plate_code(item["parent_hdx_id"][-3:]), normalize_key(item["name"]))
        district_hdx_map[key].append(item)

    district_overrides = []
    district_override_ids = set()
    for key, hdx_items in district_hdx_map.items():
        uab_names = district_uab_name_map.get(key)
        if not uab_names or len(uab_names) != 1 or len(hdx_items) != 1:
            continue

        corrected_name = next(iter(uab_names))
        hdx_item = hdx_items[0]
        if corrected_name != hdx_item["name"]:
            district_overrides.append(
                {
                    "hdx_id": hdx_item["hdx_id"],
                    "plate_code": canonical_plate_code(hdx_item["parent_hdx_id"][-3:]),
                    "name": corrected_name,
                    "source": "uab_validation_only_exact_name_match",
                }
            )
            district_override_ids.add(hdx_item["hdx_id"])

    for item in districts:
        if item["hdx_id"] in district_override_ids:
            continue

        parent = provinces_by_source.get(item["parent_hdx_id"])
        if not parent:
            continue

        if normalize_key(item["name"]) != normalize_key(parent["name"]):
            continue

        corrected_name = province_name_map.get(parent["hdx_id"], parent["name"])
        if corrected_name == item["name"]:
            continue

        district_overrides.append(
            {
                "hdx_id": item["hdx_id"],
                "plate_code": canonical_plate_code(item["parent_hdx_id"][-3:]),
                "name": corrected_name,
                "source": "derived_from_parent_province_display_name",
            }
        )
        district_override_ids.add(item["hdx_id"])

    payload = {
        "source": "UAB ilce-listesi.xlsx validation-only display name overrides",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "stats": {
            "uab_rows": len(uab_rows),
            "province_overrides": len(province_overrides),
            "district_overrides": len(district_overrides),
        },
        "province_overrides": province_overrides,
        "district_overrides": district_overrides,
    }

    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(
        f"Generated display name overrides: provinces={len(province_overrides)} districts={len(district_overrides)}"
    )


if __name__ == "__main__":
    main()
