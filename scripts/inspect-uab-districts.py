#!/usr/bin/env python

from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
UAB_PATH = ROOT / "source" / "reference" / "ilce-listesi.xlsx"
HDX_PROVINCES_PATH = ROOT / "data" / "processed" / "provinces.metadata.json"
HDX_DISTRICTS_PATH = ROOT / "data" / "processed" / "districts.metadata.json"
REPORT_PATH = ROOT / "source" / "reference" / "uab-validation-report.json"


def normalize_name(value: str) -> str:
    value = (
        str(value)
        .replace("\u0130", "i")
        .replace("I", "i")
        .replace("\u0131", "i")
        .replace("\u00c7", "c")
        .replace("\u00e7", "c")
        .replace("\u011e", "g")
        .replace("\u011f", "g")
        .replace("\u00d6", "o")
        .replace("\u00f6", "o")
        .replace("\u015e", "s")
        .replace("\u015f", "s")
        .replace("\u00dc", "u")
        .replace("\u00fc", "u")
    )
    value = unicodedata.normalize("NFKD", value)
    value = "".join(char for char in value if not unicodedata.combining(char))
    value = re.sub(r"[^a-z0-9]+", "", value.lower())
    return value


def main() -> int:
    if not UAB_PATH.exists():
        raise SystemExit(f"Missing validation source: {UAB_PATH}")

    provinces = json.loads(HDX_PROVINCES_PATH.read_text(encoding="utf-8"))
    districts = json.loads(HDX_DISTRICTS_PATH.read_text(encoding="utf-8"))
    province_by_id = {row["id"]: row for row in provinces}
    district_keys = {
        (
            row["plate_code"],
            normalize_name(row["name"]),
            normalize_name(province_by_id[row["parent_id"]]["name"]),
        )
        for row in districts
    }

    workbook = load_workbook(UAB_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))

    matched = 0
    unmatched = []
    airport_like = 0
    merkez_rows = 0

    for kodu, il_kodu, ad, il_adi in rows:
        plate_code = str(il_kodu).zfill(2)
        district_name = str(ad).strip()
        province_name = str(il_adi).strip()
        key = (plate_code, normalize_name(district_name), normalize_name(province_name))

        if "Havaliman" in district_name or "Havalimanı" in district_name:
            airport_like += 1
        if district_name.upper() == "MERKEZ":
            merkez_rows += 1

        if key in district_keys:
            matched += 1
        else:
            unmatched.append(
                {
                    "kodu": str(kodu),
                    "il_kodu": plate_code,
                    "ad": district_name,
                    "il_adi": province_name,
                }
            )

    report = {
        "source": "UAB ilce-listesi.xlsx",
        "row_count": len(rows),
        "matched_rows": matched,
        "unmatched_rows": len(unmatched),
        "airport_like_rows": airport_like,
        "merkez_rows": merkez_rows,
        "top_unmatched_provinces": Counter(row["il_adi"] for row in unmatched).most_common(20),
        "sample_unmatched_rows": unmatched[:50],
        "policy": "validation_only",
        "decision": "Do not use as canonical district crosswalk without a curated exclusion/mapping layer.",
    }

    REPORT_PATH.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {REPORT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
